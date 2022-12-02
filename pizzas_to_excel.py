import pandas as pd
import openpyxl
import json
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

orders_n = pd.read_csv('./data/orders_n.csv', sep=',', encoding='latin-1')
prediccion = pd.read_csv('./data/prediccion.csv', sep=';', encoding='latin-1')
with open('./data/ganancias_mensuales.json', 'r') as f:
    ganancias_mensuales = json.load(f)

wb = openpyxl.Workbook()

# Sheet 1: Reporte de ventas
ws = wb.active
ws.title = 'Reporte de ventas'
ws['A1'] = 'Reporte de ventas'
ws['A1'].font = Font(name='Arial', size=20, bold=True)
ws['A1'].alignment = Alignment(horizontal='center')
ws.merge_cells('A1:K1')
ws['A2'] = 'Ganancias mensuales'
ws['A2'].font = Font(name='Arial', size=14, bold=True)
ws['A2'].alignment = Alignment(horizontal='center')
ws.merge_cells('A2:K2')

meses = ganancias_mensuales.keys()
ganancias = ganancias_mensuales.values()
ganancias_mensuales = pd.DataFrame({'mes': meses, 'ganancias': ganancias})

for r in dataframe_to_rows(ganancias_mensuales, index=False, header=True):
    ws.append(r)

chart = BarChart()
chart.style = 10
data = Reference(ws, min_col=2, min_row=4, max_row=15, max_col=2)
cats = Reference(ws, min_col=1, min_row=4, max_row=15)
chart.add_data(data)
chart.set_categories(cats)
chart.legend = None
ws.add_chart(chart, 'C3')

# Sheet 2: Reporte de ingredientes
ws2 = wb.create_sheet('Reporte de ingredientes')
ws2['A1'] = 'Reporte de ingredientes'
ws2['A1'].font = Font(name='Arial', size=20, bold=True)
ws2['A1'].alignment = Alignment(horizontal='center')
ws2.merge_cells('A1:R1')
ws2['A2'] = 'Prediccion de ingredientes para la semana'
ws2['A2'].font = Font(name='Arial', size=14, bold=True)
ws2['A2'].alignment = Alignment(horizontal='center')
ws2.merge_cells('A2:R2')

for columna in prediccion.columns:
    ws2.append([columna, (prediccion[columna].sum()/53).round(2)])

img = openpyxl.drawing.image.Image('./data/ingredientes.png')
img.anchor = 'C3'
ws2.add_image(img)

# Sheet 3: Reporte de pedidos
ws3 = wb.create_sheet('Reporte de pedidos')
ws3['A1'] = 'Reporte de pedidos'
ws3['A1'].font = Font(name='Arial', size=20, bold=True)
ws3['A1'].alignment = Alignment(horizontal='center')
ws3.merge_cells('A1:K1')
ws3['A2'] = 'Numero de pedidos por dia de la semana'
ws3['A2'].font = Font(name='Arial', size=14, bold=True)
ws3['A2'].alignment = Alignment(horizontal='center')
ws3.merge_cells('A2:K2')

semana = [0,0,0,0,0,0,0]
for i,row in orders_n.iterrows():
    x = row['date']
    a = pd.to_datetime(x, format='%Y-%m-%d')
    dia = a.dayofweek
    semana[dia] += 1
    
ws3.append(['Lunes', semana[0]])
ws3.append(['Martes', semana[1]])
ws3.append(['Miercoles', semana[2]])
ws3.append(['Jueves', semana[3]])
ws3.append(['Viernes', semana[4]])
ws3.append(['Sabado', semana[5]])
ws3.append(['Domingo', semana[6]])

chart = BarChart()
chart.style = 10
data = Reference(ws3, min_col=2, min_row=4, max_row=9, max_col=2)
categorias = Reference(ws3, min_col=1, min_row=4, max_row=9)
chart.add_data(data)
chart.set_categories(categorias)
chart.legend = None
ws3.add_chart(chart, 'C3')

# exportar excel
wb.save('reporte_ventas.xlsx')

